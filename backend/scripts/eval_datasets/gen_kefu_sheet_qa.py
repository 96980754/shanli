#!/usr/bin/env python3
"""导出指定 sheet 过滤后的全量问答对（问题→解决方法，加 sheet 前缀）。

与 build_kefu_testset.py 同一过滤口径（线下操作类答案、封面/非问题行、短答案均排除），
只是不抽样、做单表全量导出——供业务方核对单表全部可测问答对，
或作为全量测试集驱动 run_agent_e2e.py。

用法（宿主机，openpyxl 读 xlsx）：
  python3 gen_kefu_sheet_qa.py \
      --source ../../../docs/【客服】POCSTARS知识库.xlsx \
      --sheet miniserver [--out miniserver_all.jsonl]
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl

from build_kefu_testset import _HEADER_LIKE, _NON_QUESTION, is_offline_answer, iter_sheet_qa

BASE = Path(__file__).resolve().parent


def iter_filtered_qa(ws) -> list[tuple[str, str]]:
    """与 build_kefu_testset 同一过滤口径：只保留可测问答对。"""
    return [
        p
        for p in iter_sheet_qa(ws)
        if not is_offline_answer(p[1]) and not _HEADER_LIKE.match(p[0]) and not _NON_QUESTION.search(p[0])
    ]


def main() -> int:
    parser = argparse.ArgumentParser(description="导出指定 sheet 的全部问答对")
    parser.add_argument("--source", required=True, help="xlsx 路径（宿主路径）")
    parser.add_argument("--sheet", required=True, help="sheet 名")
    parser.add_argument("--out", default="", help="输出 jsonl 相对 eval_datasets（默认 {sheet}_all.jsonl）")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.source, read_only=True, data_only=True)
    if args.sheet not in wb.sheetnames:
        print(f"sheet 不存在: {args.sheet}（可用: {wb.sheetnames}）")
        return 1
    pairs = iter_filtered_qa(wb[args.sheet])

    items = [
        {"index": i, "section": args.sheet, "query": f"{args.sheet}-{q}", "gold_answer": a}
        for i, (q, a) in enumerate(pairs, 1)
    ]

    out = BASE / (args.out or f"{args.sheet}_all.jsonl")
    with open(out, "w", encoding="utf-8") as f:
        for item in items:
            f.write(json.dumps(item, ensure_ascii=False) + "\n")

    print(f"[{args.sheet}] 过滤后共 {len(items)} 条问答对 -> {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
