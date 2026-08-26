#!/usr/bin/env python3
"""从甲方《【客服】POCSTARS知识库.xlsx》构建问答测试集（内部工具）。

以「问题 → 解决方法」为问答对：问题列做 query，解决方法列做 gold_answer（甲方官方标准答案），
按表（域）分层抽样 N 题，写 eval_datasets/kefu{N}.jsonl 供 run_agent_e2e.py 驱动
真实 Agent 作答。域（表名）写入 section 字段，随 run_agent_e2e 透传到结果与评分。

生成的问题统一加 sheet 前缀（如「调度台-什么是下发消息」）：携带模块上下文去问，
对齐用户实际在哪个模块提问的场景，也降低多义澄清率。

过滤规则（保证样本可测、不冤枉系统）：
- 非问答表跳过：终端知识库（型号清单）、问题等级及处理（severity 矩阵）、
  终端和调度器错误代码（错误码表）、话术（客服话术模板）、知识库（封面页）、培训知识库；
- 问答两列需非空且长度足够；
- 去除问题前缀编号（"1."、"2. "）；
- 排除「线下操作」类答案（发给/寄回/联系售后/提交工单/找销售 等）——系统无法从知识库
  执行线下动作，测它们属预期失败，不纳入本次对比。

用法（宿主机，无需容器，openpyxl 读 xlsx）：
  python3 build_kefu_testset.py \
      --source ../../../docs/【客服】POCSTARS知识库.xlsx \
      --sample 20 [--out kefu20.jsonl] [--seed 42]
"""
from __future__ import annotations

import argparse
import json
import random
import re
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent

# 域 → (问题列号0-based, 答案列号0-based, 答案别名)
SHEET_COLS: dict[str, tuple[int, int, str]] = {
    "运营平台": (1, 2, "解决方法"),
    "调度台": (1, 2, "解决方法"),
    "终端-安卓": (1, 2, "解决方法"),
    "终端-cat1": (1, 2, "解决方法"),
    "MDM": (1, 2, "解决方法"),
    "miniserver": (1, 2, "A"),
    # 测试只针对以下 6 个 sheet（业务方指令）；培训知识库不参与
}

# 非问答表（型号清单/矩阵/错误码表/话术/封面/附件索引），跳过
SKIP_SHEETS = {
    "终端知识库",
    "问题等级及处理",
    "终端和调度器错误代码",
    "话术",
    "知识库",
    "快速部署",
    "流量卡问题",
}

# 线下操作类答案关键词：系统无法从知识库执行线下动作，测它们属预期失败
_OFFLINE_OPS = (
    "发给客户", "寄回", "联系售后", "提交工单", "找销售", "联系销售", "线下",
    "放到qt下", "发到qt", "走售后",
)

# 封面/目录词（须在句首）
_HEADER_LIKE = re.compile(r"^(第[一二三四五六七八九十\d]+章|目录|前言|概述|说明|备注|注:|注意)")
# 非问题行模式（出现在任意位置）：文件名/媒体资产/发布说明/截图请求
_NON_QUESTION = re.compile(
    r"\.(docx|xlsx|doc|xls|pdf|csv|txt|zip|mp4|jpg|png|jpeg|gif)$"
    r"|^[Qq][1-4]\s*发布"  # 发布说明（如"Q3发布派接组功能"）
    r"|(页面截图|界面截图)$"  # 要截图的资产请求（图片流程类）
)


def clean_question(text: str) -> str:
    """去编号前缀、折叠空白。"""
    text = re.sub(r"^\s*\d+[\.、）)]\s*", "", str(text).strip())
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def iter_sheet_qa(ws) -> list[tuple[str, str]]:
    """按表头定位问题/答案列，返回该表 (问题, 解决方法) 列表。"""
    rows = list(ws.iter_rows(values_only=True))
    q_col, a_col, a_alias = SHEET_COLS.get(ws.title, (1, 2, "解决方法"))
    qa: list[tuple[str, str]] = []
    started = False
    for row in rows:
        if not row or not any(c is not None for c in row):
            continue
        q_cell = row[q_col] if q_col < len(row) else None
        a_cell = row[a_col] if a_col < len(row) else None
        if not started:
            # 表头行：问题列出现「问题」或 miniserver 的「Q」
            if q_cell is not None and (str(q_cell).strip() in ("问题", "Q") or "问题" in str(q_cell)):
                started = True
            continue
        q = clean_question(q_cell or "")
        a = clean_question(a_cell or "")
        if len(q) < 4 or len(a) < 4:
            continue
        qa.append((q, a))
    return qa


def is_offline_answer(answer: str) -> bool:
    return any(kw in answer for kw in _OFFLINE_OPS)


def main() -> int:
    parser = argparse.ArgumentParser(description="从甲方客服知识库 xlsx 构建问答测试集")
    parser.add_argument("--source", required=True, help="xlsx 路径（宿主路径）")
    parser.add_argument("--sample", type=int, default=20, help="每轮抽样总题数（0=全部）")
    parser.add_argument("--out", default="", help="输出 jsonl 相对 eval_datasets（默认 synthetic/kefu{N}.jsonl）")
    parser.add_argument("--seed", type=int, default=42)
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.source, read_only=True, data_only=True)
    domain_rows: dict[str, list[tuple[str, str]]] = {}
    for name in SHEET_COLS:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        pairs = iter_sheet_qa(ws)
        kept = [
            p
            for p in pairs
            if not is_offline_answer(p[1]) and not _HEADER_LIKE.match(p[0]) and not _NON_QUESTION.search(p[0])
        ]
        if kept:
            domain_rows[name] = kept
        print(f"  {name}: {len(pairs)} 问答对（过滤后 {len(kept)}）")

    # 按域分配名额：优先均匀覆盖，余数给题量大的域
    order = sorted(domain_rows, key=lambda k: -len(domain_rows[k]))
    slots: dict[str, int] = {}
    if args.sample:
        base, rem = divmod(args.sample, len(order))
        for i, name in enumerate(order):
            slots[name] = base + (1 if i < rem else 0)
    else:
        for name in order:
            slots[name] = len(domain_rows[name])

    rng = random.Random(args.seed)
    chosen: list[dict] = []
    index = 1
    for name in order:
        pool = domain_rows[name]
        n = min(slots[name], len(pool))
        for q, a in rng.sample(pool, n):
            # 问题加 sheet 前缀（业务方指令）：携带模块上下文去问，降低多义澄清率
            chosen.append({"index": index, "section": name, "query": f"{name}-{q}", "gold_answer": a})
            index += 1

    # 默认写在 eval_datasets/ 下（synthetic/ 为容器 root 所有，宿主不可写）
    out_rel = args.out or f"kefu{args.sample or 'all'}.jsonl"
    out = BASE / out_rel
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8") as f:
        for item in chosen:
            f.write(json.dumps(item, ensure_ascii=False) + "\n")

    print(f"\n共抽 {len(chosen)} 题 -> {out}")
    for item in chosen:
        print(f"  #{item['index']:>2} [{item['section']}] {item['query'][:44]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
