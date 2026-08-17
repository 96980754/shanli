from __future__ import annotations

from collections import Counter
from io import BytesIO
from pathlib import PurePath
from typing import Any
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from yuxi.services.user_identity_service import (
    is_valid_phone_number,
    normalize_phone_number,
    validate_explicit_uid,
    validate_username,
)
from yuxi.storage.postgres.models_business import Department, User
from yuxi.utils.auth_utils import AuthUtils

USER_IMPORT_SHEET = "用户导入"
USER_IMPORT_HEADERS = (
    "username",
    "uid",
    "initial_password",
    "phone_number",
    "role",
    "department_name",
)
MAX_USER_IMPORT_BYTES = 2 * 1024 * 1024
MAX_USER_IMPORT_ROWS = 200
MAX_USER_IMPORT_ERRORS = 500
MAX_XLSX_EXPANDED_BYTES = 20 * 1024 * 1024


class UserImportFileError(ValueError):
    pass


def build_user_import_template(*, departments: list[str], is_superadmin: bool) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = USER_IMPORT_SHEET
    sheet.append(USER_IMPORT_HEADERS)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:F1"
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 26
    sheet.column_dimensions["C"].width = 24
    sheet.column_dimensions["D"].width = 18
    sheet.column_dimensions["E"].width = 14
    sheet.column_dimensions["F"].width = 22

    required_fill = PatternFill("solid", fgColor="DCEBFF")
    optional_fill = PatternFill("solid", fgColor="F2F2F2")
    for index, cell in enumerate(sheet[1], start=1):
        cell.font = Font(bold=True)
        cell.fill = optional_fill if index in (4, 5) else required_fill

    for column in ("B", "C", "D"):
        for row in range(2, MAX_USER_IMPORT_ROWS + 2):
            sheet[f"{column}{row}"].number_format = "@"

    role_validation = DataValidation(
        type="list",
        formula1='"user,admin"' if is_superadmin else '"user"',
        allow_blank=True,
    )
    sheet.add_data_validation(role_validation)
    role_validation.add(f"E2:E{MAX_USER_IMPORT_ROWS + 1}")

    if is_superadmin and departments:
        options = workbook.create_sheet("选项")
        options.sheet_state = "hidden"
        for index, name in enumerate(departments, start=1):
            options.cell(index, 1, name)
        department_validation = DataValidation(
            type="list",
            formula1=f"'选项'!$A$1:$A${len(departments)}",
            allow_blank=False,
        )
        sheet.add_data_validation(department_validation)
        department_validation.add(f"F2:F{MAX_USER_IMPORT_ROWS + 1}")

    instructions = workbook.create_sheet("填写说明")
    instructions.append(["字段", "要求"])
    instructions.append(["username", "必填，2-20个字符，只能包含中文、英文、数字和下划线。"])
    instructions.append(["uid", "必填且导入后不可修改；企业微信用户需精确填写企微 UserID。"])
    instructions.append(["initial_password", "必填，8-128个字符；文件包含明文密码，导入后请安全删除。"])
    instructions.append(["phone_number", "选填，中国大陆手机号。"])
    instructions.append(["role", "选填，留空默认 user；仅允许 user 或 admin，部门管理员只能导入 user。"])
    instructions.append(["department_name", "超级管理员导入时必填；部门管理员只能使用自己的部门。"])
    instructions.column_dimensions["A"].width = 24
    instructions.column_dimensions["B"].width = 90
    instructions[1][0].font = instructions[1][1].font = Font(bold=True)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _validate_xlsx_archive(data: bytes) -> None:
    try:
        with ZipFile(BytesIO(data)) as archive:
            total_size = sum(info.file_size for info in archive.infolist())
    except BadZipFile as exc:
        raise UserImportFileError("文件不是有效的 XLSX 工作簿") from exc
    if total_size > MAX_XLSX_EXPANDED_BYTES:
        raise UserImportFileError("XLSX 解压后内容过大")


def _cell_text(cell, *, field: str, row_number: int, errors: list[dict[str, Any]]) -> str:
    if cell.data_type == "f":
        errors.append(_error(row_number, field, "formula_not_allowed", f"{field} 不允许使用公式"))
        return ""
    if cell.value is None:
        return ""
    if not isinstance(cell.value, str):
        errors.append(_error(row_number, field, "invalid_cell_type", f"{field} 必须是文本格式"))
        return ""
    return cell.value


def _error(row: int, column: str, code: str, message: str) -> dict[str, Any]:
    return {"excel_row": row, "column": column, "code": code, "message": message}


def parse_user_import_workbook(data: bytes) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    _validate_xlsx_archive(data)
    try:
        workbook = load_workbook(BytesIO(data), read_only=False, data_only=False)
    except Exception as exc:
        raise UserImportFileError("无法读取 XLSX 工作簿") from exc
    if USER_IMPORT_SHEET not in workbook.sheetnames:
        raise UserImportFileError(f"工作簿必须包含“{USER_IMPORT_SHEET}”工作表")

    sheet = workbook[USER_IMPORT_SHEET]
    headers = tuple(cell.value for cell in sheet[1])
    if headers != USER_IMPORT_HEADERS:
        raise UserImportFileError(f"表头必须严格为：{', '.join(USER_IMPORT_HEADERS)}")
    if sheet.merged_cells.ranges:
        raise UserImportFileError("用户导入工作表不允许包含合并单元格")

    rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    for excel_row, cells in enumerate(sheet.iter_rows(min_row=2, max_col=len(USER_IMPORT_HEADERS)), start=2):
        if all(cell.value is None or (isinstance(cell.value, str) and not cell.value.strip()) for cell in cells):
            continue
        if len(rows) >= MAX_USER_IMPORT_ROWS:
            raise UserImportFileError(f"单次最多导入 {MAX_USER_IMPORT_ROWS} 个用户")

        values = {
            field: _cell_text(cell, field=field, row_number=excel_row, errors=errors)
            for field, cell in zip(USER_IMPORT_HEADERS, cells, strict=True)
        }
        username = values["username"].strip()
        uid = values["uid"]
        password = values["initial_password"]
        phone = normalize_phone_number(values["phone_number"]) if values["phone_number"] else ""
        role = values["role"].strip() or "user"
        department_name = values["department_name"].strip()

        valid_username, username_error = validate_username(username)
        if not valid_username:
            errors.append(_error(excel_row, "username", "invalid_username", username_error))
        valid_uid, uid_error = validate_explicit_uid(uid)
        if not valid_uid:
            errors.append(_error(excel_row, "uid", "invalid_uid", uid_error))
        if not isinstance(password, str) or len(password) < 8 or len(password) > 128 or not password.strip():
            errors.append(_error(excel_row, "initial_password", "invalid_password", "初始密码长度必须为8-128个字符"))
        if phone and not is_valid_phone_number(phone):
            errors.append(_error(excel_row, "phone_number", "invalid_phone", "手机号格式不正确"))
        if role not in {"user", "admin"}:
            errors.append(_error(excel_row, "role", "invalid_role", "角色仅允许 user 或 admin，留空默认为 user"))

        rows.append(
            {
                "excel_row": excel_row,
                "username": username,
                "uid": uid,
                "initial_password": password,
                "phone_number": phone or None,
                "role": role,
                "department_name": department_name,
            }
        )

    if not rows:
        raise UserImportFileError("用户导入工作表没有数据")
    return rows, errors[:MAX_USER_IMPORT_ERRORS]


def _append_duplicate_errors(rows: list[dict[str, Any]], errors: list[dict[str, Any]]) -> None:
    for field in ("username", "uid", "phone_number"):
        locations: dict[str, list[int]] = {}
        for row in rows:
            value = row[field]
            if value:
                locations.setdefault(value, []).append(row["excel_row"])
        for row_numbers in locations.values():
            if len(row_numbers) < 2:
                continue
            first = row_numbers[0]
            for row_number in row_numbers:
                errors.append(_error(row_number, field, "duplicate_in_file", f"与 Excel 第 {first} 行重复"))


async def validate_user_import(
    data: bytes,
    *,
    current_user: User,
    session: AsyncSession,
) -> dict[str, Any]:
    rows, errors = parse_user_import_workbook(data)
    _append_duplicate_errors(rows, errors)

    department_result = await session.execute(select(Department))
    departments = list(department_result.scalars().all())
    departments_by_name = {department.name: department for department in departments}
    current_department = next((item for item in departments if item.id == current_user.department_id), None)

    for row in rows:
        if current_user.role == "superadmin":
            if not row["department_name"]:
                errors.append(_error(row["excel_row"], "department_name", "required", "超级管理员导入时必须填写部门"))
            elif row["department_name"] not in departments_by_name:
                errors.append(_error(row["excel_row"], "department_name", "unknown_department", "部门不存在"))
            else:
                row["department_id"] = departments_by_name[row["department_name"]].id
        else:
            if current_department is None:
                errors.append(_error(row["excel_row"], "department_name", "missing_scope", "当前管理员未绑定部门"))
            elif row["role"] != "user":
                errors.append(_error(row["excel_row"], "role", "forbidden_role", "部门管理员只能导入普通用户"))
            elif row["department_name"] and row["department_name"] != current_department.name:
                errors.append(
                    _error(row["excel_row"], "department_name", "forbidden_department", "只能导入当前部门用户")
                )
            else:
                row["department_name"] = current_department.name
                row["department_id"] = current_department.id

    usernames = {row["username"] for row in rows}
    uids = {row["uid"] for row in rows}
    phones = {row["phone_number"] for row in rows if row["phone_number"]}
    conflict_conditions = [User.username.in_(usernames), User.uid.in_(uids)]
    if phones:
        conflict_conditions.append(User.phone_number.in_(phones))
    result = await session.execute(select(User).where(or_(*conflict_conditions)))
    conflicts = list(result.scalars().all())
    conflict_usernames = {item.username for item in conflicts}
    conflict_uids = {item.uid for item in conflicts}
    conflict_phones = {item.phone_number for item in conflicts if item.phone_number}
    for row in rows:
        if row["username"] in conflict_usernames:
            errors.append(_error(row["excel_row"], "username", "already_exists", "用户名已存在"))
        if row["uid"] in conflict_uids:
            errors.append(_error(row["excel_row"], "uid", "already_exists", "UID 已存在"))
        if row["phone_number"] and row["phone_number"] in conflict_phones:
            errors.append(_error(row["excel_row"], "phone_number", "already_exists", "手机号已存在"))

    limited_errors = errors[:MAX_USER_IMPORT_ERRORS]
    preview_rows = [
        {
            "excel_row": row["excel_row"],
            "username": row["username"],
            "uid": row["uid"],
            "phone_number": _mask_phone(row["phone_number"]),
            "role": row["role"],
            "department_name": row.get("department_name", ""),
        }
        for row in rows
    ]
    return {
        "valid": not errors,
        "row_count": len(rows),
        "rows": preview_rows,
        "errors": limited_errors,
        "errors_truncated": len(errors) > len(limited_errors),
        "_rows": rows,
    }


def _mask_phone(phone: str | None) -> str | None:
    if not phone:
        return None
    return f"{phone[:3]}****{phone[-4:]}"


async def import_users_atomically(
    validation: dict[str, Any],
    *,
    session: AsyncSession,
) -> dict[str, Any]:
    if not validation["valid"]:
        raise ValueError("用户导入校验未通过")

    rows = validation["_rows"]
    role_counts = Counter(row["role"] for row in rows)
    department_names = sorted({row["department_name"] for row in rows})
    users = [
        User(
            username=row["username"],
            uid=row["uid"],
            phone_number=row["phone_number"],
            password_hash=AuthUtils.hash_password(row["initial_password"]),
            role=row["role"],
            department_id=row["department_id"],
        )
        for row in rows
    ]
    session.add_all(users)
    try:
        await session.flush()
        await session.commit()
    except Exception:
        await session.rollback()
        raise
    finally:
        for row in rows:
            row.pop("initial_password", None)

    return {
        "success": True,
        "imported_count": len(users),
        "role_counts": dict(role_counts),
        "departments": department_names,
    }


def safe_upload_filename(filename: str | None) -> str:
    return PurePath(filename or "用户导入.xlsx").name
