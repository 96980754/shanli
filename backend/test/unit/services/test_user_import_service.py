from io import BytesIO
from types import SimpleNamespace

import pytest
from openpyxl import Workbook, load_workbook

from yuxi.services.user_import_service import (
    USER_IMPORT_HEADERS,
    UserImportFileError,
    build_user_import_template,
    parse_user_import_workbook,
    validate_user_import,
)


def _workbook_bytes(rows: list[list[object]], *, headers=USER_IMPORT_HEADERS) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "用户导入"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_template_defaults_role_to_optional_user():
    data = build_user_import_template(departments=["研发部"], is_superadmin=True)
    workbook = load_workbook(BytesIO(data))

    sheet = workbook["用户导入"]
    assert tuple(cell.value for cell in sheet[1]) == USER_IMPORT_HEADERS
    assert sheet["E2"].value is None
    assert "留空默认 user" in workbook["填写说明"][6][1].value


def test_parse_defaults_blank_role_to_user_and_keeps_uid_case():
    rows, errors = parse_user_import_workbook(
        _workbook_bytes([["张三", "Zhang.San@WeCom", "password123", "13800138000", "", "研发部"]])
    )

    assert errors == []
    assert rows[0]["uid"] == "Zhang.San@WeCom"
    assert rows[0]["role"] == "user"
    assert rows[0]["phone_number"] == "13800138000"


def test_parse_rejects_formula_and_does_not_expose_password():
    rows, errors = parse_user_import_workbook(
        _workbook_bytes([["张三", "zhangsan", "=CONCAT(\"pass\",\"word\")", "", "", "研发部"]])
    )

    assert rows[0]["initial_password"] == ""
    assert any(error["column"] == "initial_password" for error in errors)
    assert all("CONCAT" not in error["message"] for error in errors)


def test_parse_requires_exact_headers():
    with pytest.raises(UserImportFileError, match="表头必须严格"):
        parse_user_import_workbook(_workbook_bytes([], headers=tuple(reversed(USER_IMPORT_HEADERS))))


@pytest.mark.asyncio
async def test_validation_rejects_duplicates_without_returning_password():
    class Scalars:
        def __init__(self, values):
            self.values = values

        def all(self):
            return self.values

    class Result:
        def __init__(self, values):
            self.values = values

        def scalars(self):
            return Scalars(self.values)

    class Session:
        calls = 0

        async def execute(self, _statement):
            self.calls += 1
            if self.calls == 1:
                return Result([SimpleNamespace(id=1, name="研发部")])
            return Result([])

    data = _workbook_bytes(
        [
            ["张三", "zhangsan", "password123", "13800138000", "", "研发部"],
            ["张三", "lisi", "anotherpass", "13900139000", "user", "研发部"],
        ]
    )
    result = await validate_user_import(
        data,
        current_user=SimpleNamespace(role="superadmin", department_id=1),
        session=Session(),
    )

    assert result["valid"] is False
    assert any(error["code"] == "duplicate_in_file" for error in result["errors"])
    assert all("initial_password" not in row for row in result["rows"])
    assert "password123" not in str({key: value for key, value in result.items() if not key.startswith("_")})


@pytest.mark.asyncio
async def test_import_rolls_back_every_user_on_commit_failure():
    from yuxi.services.user_import_service import import_users_atomically

    class Session:
        added = []
        rolled_back = False

        def add_all(self, users):
            self.added = users

        async def flush(self):
            return None

        async def commit(self):
            raise RuntimeError("commit failed")

        async def rollback(self):
            self.rolled_back = True

    rows = [
        {
            "excel_row": 2,
            "username": "张三",
            "uid": "zhangsan",
            "initial_password": "password123",
            "phone_number": None,
            "role": "user",
            "department_name": "研发部",
            "department_id": 1,
        }
    ]
    session = Session()

    with pytest.raises(RuntimeError, match="commit failed"):
        await import_users_atomically({"valid": True, "_rows": rows}, session=session)

    assert session.rolled_back is True
    assert "initial_password" not in rows[0]
    assert len(session.added) == 1
